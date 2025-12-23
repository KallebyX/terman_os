# -*- coding: utf-8 -*-
"""
Serviço de Excel
Importação e Exportação de dados em formato Excel
"""

import io
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import logging

logger = logging.getLogger(__name__)


class ExcelService:
    """Serviço para manipulação de arquivos Excel"""

    # Estilos padrão
    HEADER_FILL = PatternFill(start_color='0071E3', end_color='0071E3', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ALT_ROW_FILL = PatternFill(start_color='F5F5F7', end_color='F5F5F7', fill_type='solid')

    BORDER_STYLE = Border(
        left=Side(style='thin', color='D2D2D7'),
        right=Side(style='thin', color='D2D2D7'),
        top=Side(style='thin', color='D2D2D7'),
        bottom=Side(style='thin', color='D2D2D7')
    )

    CURRENCY_FORMAT = 'R$ #,##0.00'
    DATE_FORMAT = 'DD/MM/YYYY'
    DATETIME_FORMAT = 'DD/MM/YYYY HH:MM'
    NUMBER_FORMAT = '#,##0.00'

    def __init__(self, empresa=None):
        """
        Inicializa o serviço de Excel

        Args:
            empresa: Objeto ConfiguracaoEmpresa para cabeçalho
        """
        self.empresa = empresa

    def exportar_vendas(self, vendas, periodo=None):
        """
        Exporta relatório de vendas para Excel

        Args:
            vendas: Lista de objetos Pedido ou dicts
            periodo: String do período (opcional)

        Returns:
            bytes: Arquivo Excel em bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Vendas'

        # Cabeçalho do relatório
        self._adicionar_cabecalho_empresa(ws)

        # Título
        ws.merge_cells('A5:J5')
        title_cell = ws['A5']
        title_cell.value = f"RELATÓRIO DE VENDAS{' - ' + periodo if periodo else ''}"
        title_cell.font = Font(bold=True, size=14, color='1D1D1F')
        title_cell.alignment = Alignment(horizontal='center')

        # Cabeçalhos da tabela
        headers = ['Nº Pedido', 'Data', 'Cliente', 'Email', 'Produtos',
                   'Subtotal', 'Desconto', 'Frete', 'Total', 'Status']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_STYLE

        # Dados
        row = 8
        total_geral = 0

        for venda in vendas:
            if hasattr(venda, 'numero_pedido'):
                # Objeto Pedido
                dados = [
                    venda.numero_pedido,
                    venda.data_criacao.strftime('%d/%m/%Y') if venda.data_criacao else '',
                    venda.usuario.nome if venda.usuario else '',
                    venda.usuario.email if venda.usuario else '',
                    len(list(venda.itens)) if venda.itens else 0,
                    float(venda.subtotal or 0),
                    float(venda.desconto or 0),
                    float(venda.valor_frete or 0),
                    float(venda.total or 0),
                    venda.status
                ]
            else:
                # Dict
                dados = [
                    venda.get('numero', ''),
                    venda.get('data', ''),
                    venda.get('cliente', ''),
                    venda.get('email', ''),
                    venda.get('qtd_produtos', 0),
                    float(venda.get('subtotal', 0)),
                    float(venda.get('desconto', 0)),
                    float(venda.get('frete', 0)),
                    float(venda.get('total', 0)),
                    venda.get('status', '')
                ]

            total_geral += dados[8]

            for col, valor in enumerate(dados, 1):
                cell = ws.cell(row=row, column=col, value=valor)
                cell.border = self.BORDER_STYLE

                if col in [6, 7, 8, 9]:  # Valores monetários
                    cell.number_format = self.CURRENCY_FORMAT

                if row % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

            row += 1

        # Total geral
        row += 1
        ws.cell(row=row, column=8, value='TOTAL GERAL:').font = Font(bold=True)
        total_cell = ws.cell(row=row, column=9, value=total_geral)
        total_cell.font = Font(bold=True, color='0071E3')
        total_cell.number_format = self.CURRENCY_FORMAT

        # Ajusta larguras
        self._ajustar_larguras(ws, headers)

        # Salva em bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def exportar_produtos(self, produtos):
        """
        Exporta catálogo de produtos para Excel

        Args:
            produtos: Lista de objetos Produto ou dicts

        Returns:
            bytes: Arquivo Excel em bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Produtos'

        # Cabeçalho
        self._adicionar_cabecalho_empresa(ws)

        ws.merge_cells('A5:K5')
        title_cell = ws['A5']
        title_cell.value = "CATÁLOGO DE PRODUTOS"
        title_cell.font = Font(bold=True, size=14, color='1D1D1F')
        title_cell.alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Código', 'Nome', 'Categoria', 'Preço Custo', 'Preço Venda',
                   'Preço Promo', 'Estoque', 'Mínimo', 'NCM', 'Ativo', 'Destaque']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_STYLE

        # Dados
        row = 8
        for produto in produtos:
            if hasattr(produto, 'codigo'):
                dados = [
                    produto.codigo,
                    produto.nome,
                    produto.categoria.nome if produto.categoria else '',
                    float(produto.preco_custo or 0),
                    float(produto.preco or 0),
                    float(produto.preco_promocional or 0),
                    float(produto.estoque_total if hasattr(produto, 'estoque_total') else 0),
                    0,
                    produto.ncm or '',
                    'Sim' if produto.ativo else 'Não',
                    'Sim' if produto.destaque else 'Não'
                ]
            else:
                dados = [
                    produto.get('codigo', ''),
                    produto.get('nome', ''),
                    produto.get('categoria', ''),
                    float(produto.get('preco_custo', 0)),
                    float(produto.get('preco', 0)),
                    float(produto.get('preco_promo', 0)),
                    float(produto.get('estoque', 0)),
                    float(produto.get('minimo', 0)),
                    produto.get('ncm', ''),
                    'Sim' if produto.get('ativo') else 'Não',
                    'Sim' if produto.get('destaque') else 'Não'
                ]

            for col, valor in enumerate(dados, 1):
                cell = ws.cell(row=row, column=col, value=valor)
                cell.border = self.BORDER_STYLE

                if col in [4, 5, 6]:  # Preços
                    cell.number_format = self.CURRENCY_FORMAT

                if row % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

            row += 1

        self._ajustar_larguras(ws, headers)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def exportar_estoque(self, estoques):
        """
        Exporta relatório de estoque para Excel

        Args:
            estoques: Lista de objetos Estoque ou dicts

        Returns:
            bytes: Arquivo Excel em bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Estoque'

        self._adicionar_cabecalho_empresa(ws)

        ws.merge_cells('A5:I5')
        title_cell = ws['A5']
        title_cell.value = "RELATÓRIO DE ESTOQUE"
        title_cell.font = Font(bold=True, size=14, color='1D1D1F')
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['Código', 'Produto', 'Localização', 'Quantidade',
                   'Mínimo', 'Máximo', 'Status', 'Custo Unit', 'Valor Total']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_STYLE

        row = 8
        valor_total_geral = 0

        for estoque in estoques:
            if hasattr(estoque, 'produto'):
                qtd = float(estoque.quantidade or 0)
                custo = float(estoque.produto.preco_custo or 0) if estoque.produto else 0
                valor_total = qtd * custo
                valor_total_geral += valor_total

                status = 'OK'
                if qtd <= 0:
                    status = 'SEM ESTOQUE'
                elif qtd <= float(estoque.quantidade_minima or 0):
                    status = 'BAIXO'
                elif qtd >= float(estoque.quantidade_maxima or 999999):
                    status = 'EXCESSO'

                dados = [
                    estoque.produto.codigo if estoque.produto else '',
                    estoque.produto.nome if estoque.produto else '',
                    estoque.localizacao or 'Principal',
                    qtd,
                    float(estoque.quantidade_minima or 0),
                    float(estoque.quantidade_maxima or 0),
                    status,
                    custo,
                    valor_total
                ]
            else:
                qtd = float(estoque.get('quantidade', 0))
                custo = float(estoque.get('custo', 0))
                valor_total = qtd * custo
                valor_total_geral += valor_total

                dados = [
                    estoque.get('codigo', ''),
                    estoque.get('produto', ''),
                    estoque.get('localizacao', 'Principal'),
                    qtd,
                    float(estoque.get('minimo', 0)),
                    float(estoque.get('maximo', 0)),
                    estoque.get('status', 'OK'),
                    custo,
                    valor_total
                ]

            for col, valor in enumerate(dados, 1):
                cell = ws.cell(row=row, column=col, value=valor)
                cell.border = self.BORDER_STYLE

                if col in [8, 9]:
                    cell.number_format = self.CURRENCY_FORMAT

                # Destaca status
                if col == 7:
                    if valor == 'SEM ESTOQUE':
                        cell.font = Font(color='FF453A', bold=True)
                    elif valor == 'BAIXO':
                        cell.font = Font(color='FF9F0A', bold=True)

                if row % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

            row += 1

        # Total
        row += 1
        ws.cell(row=row, column=8, value='VALOR TOTAL:').font = Font(bold=True)
        total_cell = ws.cell(row=row, column=9, value=valor_total_geral)
        total_cell.font = Font(bold=True, color='0071E3')
        total_cell.number_format = self.CURRENCY_FORMAT

        self._ajustar_larguras(ws, headers)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def exportar_clientes(self, clientes):
        """
        Exporta lista de clientes para Excel

        Args:
            clientes: Lista de objetos Cliente ou dicts

        Returns:
            bytes: Arquivo Excel em bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clientes'

        self._adicionar_cabecalho_empresa(ws)

        ws.merge_cells('A5:L5')
        title_cell = ws['A5']
        title_cell.value = "CADASTRO DE CLIENTES"
        title_cell.font = Font(bold=True, size=14, color='1D1D1F')
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['CPF/CNPJ', 'Nome', 'Email', 'Telefone', 'Cidade', 'UF',
                   'Tipo', 'Categoria', 'Total Compras', 'Qtd Pedidos',
                   'Última Compra', 'Status']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_STYLE

        row = 8
        for cliente in clientes:
            if hasattr(cliente, 'cpf_cnpj'):
                dados = [
                    cliente.cpf_cnpj,
                    cliente.usuario.nome if cliente.usuario else '',
                    cliente.usuario.email if cliente.usuario else '',
                    cliente.telefone or cliente.celular or '',
                    cliente.cidade or '',
                    cliente.estado or '',
                    cliente.tipo or 'varejo',
                    cliente.categoria or 'Regular',
                    float(cliente.total_compras or 0),
                    cliente.quantidade_pedidos or 0,
                    cliente.ultima_compra.strftime('%d/%m/%Y') if cliente.ultima_compra else '',
                    'Ativo' if cliente.ativo else 'Inativo'
                ]
            else:
                dados = [
                    cliente.get('cpf_cnpj', ''),
                    cliente.get('nome', ''),
                    cliente.get('email', ''),
                    cliente.get('telefone', ''),
                    cliente.get('cidade', ''),
                    cliente.get('uf', ''),
                    cliente.get('tipo', 'varejo'),
                    cliente.get('categoria', 'Regular'),
                    float(cliente.get('total_compras', 0)),
                    cliente.get('qtd_pedidos', 0),
                    cliente.get('ultima_compra', ''),
                    cliente.get('status', 'Ativo')
                ]

            for col, valor in enumerate(dados, 1):
                cell = ws.cell(row=row, column=col, value=valor)
                cell.border = self.BORDER_STYLE

                if col == 9:
                    cell.number_format = self.CURRENCY_FORMAT

                if row % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

            row += 1

        self._ajustar_larguras(ws, headers)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def exportar_nfes(self, notas):
        """
        Exporta lista de NFes para Excel

        Args:
            notas: Lista de objetos NotaFiscal

        Returns:
            bytes: Arquivo Excel em bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'NFes'

        self._adicionar_cabecalho_empresa(ws)

        ws.merge_cells('A5:K5')
        title_cell = ws['A5']
        title_cell.value = "NOTAS FISCAIS ELETRÔNICAS"
        title_cell.font = Font(bold=True, size=14, color='1D1D1F')
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['Número', 'Série', 'Data Emissão', 'Destinatário', 'CNPJ/CPF',
                   'Valor Produtos', 'ICMS', 'Total', 'Status', 'Protocolo', 'Chave']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_STYLE

        row = 8
        total_geral = 0
        total_icms = 0

        for nfe in notas:
            total_geral += float(nfe.valor_total or 0)
            total_icms += float(nfe.valor_icms or 0)

            dados = [
                nfe.numero,
                nfe.serie,
                nfe.data_emissao.strftime('%d/%m/%Y') if nfe.data_emissao else '',
                nfe.destinatario_razao_social or '',
                nfe.destinatario_cpf_cnpj or '',
                float(nfe.valor_produtos or 0),
                float(nfe.valor_icms or 0),
                float(nfe.valor_total or 0),
                nfe.status,
                nfe.protocolo_autorizacao or '',
                nfe.chave_acesso or ''
            ]

            for col, valor in enumerate(dados, 1):
                cell = ws.cell(row=row, column=col, value=valor)
                cell.border = self.BORDER_STYLE

                if col in [6, 7, 8]:
                    cell.number_format = self.CURRENCY_FORMAT

                if row % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL

            row += 1

        # Totais
        row += 1
        ws.cell(row=row, column=6, value=total_geral).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=7, value=total_icms).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=8, value=total_geral).number_format = self.CURRENCY_FORMAT

        for col in [6, 7, 8]:
            ws.cell(row=row, column=col).font = Font(bold=True)

        self._ajustar_larguras(ws, headers)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def importar_produtos(self, arquivo):
        """
        Importa produtos de arquivo Excel

        Args:
            arquivo: Arquivo Excel (bytes ou file-like)

        Returns:
            dict: Resultado da importação com lista de produtos
        """
        try:
            if isinstance(arquivo, bytes):
                arquivo = io.BytesIO(arquivo)

            wb = openpyxl.load_workbook(arquivo)
            ws = wb.active

            # Encontra linha de cabeçalho
            header_row = None
            for row in range(1, 20):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and str(cell_value).lower() in ['codigo', 'código', 'sku', 'cod']:
                    header_row = row
                    break

            if not header_row:
                return {
                    'sucesso': False,
                    'erro': 'Cabeçalho não encontrado. A primeira coluna deve conter "Código" ou "SKU"'
                }

            # Mapeia colunas
            headers = {}
            for col in range(1, ws.max_column + 1):
                header = str(ws.cell(row=header_row, column=col).value or '').lower().strip()
                headers[header] = col

            # Mapeamento de campos
            campo_map = {
                'codigo': ['codigo', 'código', 'sku', 'cod'],
                'nome': ['nome', 'produto', 'descricao', 'descrição', 'name'],
                'preco': ['preco', 'preço', 'valor', 'price', 'preco_venda'],
                'preco_custo': ['preco_custo', 'custo', 'cost'],
                'preco_promocional': ['preco_promo', 'promocional', 'promo'],
                'categoria': ['categoria', 'category'],
                'ncm': ['ncm'],
                'unidade': ['unidade', 'un', 'unit'],
                'estoque': ['estoque', 'quantidade', 'qtd', 'stock'],
                'descricao': ['descricao_longa', 'descricao_completa', 'description'],
            }

            def get_col(campos):
                for campo in campos:
                    if campo in headers:
                        return headers[campo]
                return None

            produtos = []
            erros = []

            for row in range(header_row + 1, ws.max_row + 1):
                # Pula linhas vazias
                codigo = ws.cell(row=row, column=get_col(campo_map['codigo']) or 1).value
                if not codigo:
                    continue

                try:
                    produto = {
                        'codigo': str(codigo).strip(),
                        'nome': str(ws.cell(row=row, column=get_col(campo_map['nome']) or 2).value or '').strip(),
                        'preco': self._parse_decimal(ws.cell(row=row, column=get_col(campo_map['preco']) or 3).value),
                        'preco_custo': self._parse_decimal(ws.cell(row=row, column=get_col(campo_map['preco_custo'])).value) if get_col(campo_map['preco_custo']) else 0,
                        'preco_promocional': self._parse_decimal(ws.cell(row=row, column=get_col(campo_map['preco_promocional'])).value) if get_col(campo_map['preco_promocional']) else None,
                        'categoria': str(ws.cell(row=row, column=get_col(campo_map['categoria'])).value or '').strip() if get_col(campo_map['categoria']) else '',
                        'ncm': str(ws.cell(row=row, column=get_col(campo_map['ncm'])).value or '').strip() if get_col(campo_map['ncm']) else '',
                        'unidade': str(ws.cell(row=row, column=get_col(campo_map['unidade'])).value or 'UN').strip().upper() if get_col(campo_map['unidade']) else 'UN',
                        'estoque': self._parse_decimal(ws.cell(row=row, column=get_col(campo_map['estoque'])).value) if get_col(campo_map['estoque']) else 0,
                        'linha': row
                    }

                    if not produto['nome']:
                        erros.append(f"Linha {row}: Nome do produto é obrigatório")
                        continue

                    produtos.append(produto)

                except Exception as e:
                    erros.append(f"Linha {row}: {str(e)}")

            return {
                'sucesso': True,
                'produtos': produtos,
                'total': len(produtos),
                'erros': erros
            }

        except Exception as e:
            logger.error(f"Erro ao importar produtos: {str(e)}")
            return {
                'sucesso': False,
                'erro': str(e)
            }

    def importar_clientes(self, arquivo):
        """
        Importa clientes de arquivo Excel

        Args:
            arquivo: Arquivo Excel (bytes ou file-like)

        Returns:
            dict: Resultado da importação
        """
        try:
            if isinstance(arquivo, bytes):
                arquivo = io.BytesIO(arquivo)

            wb = openpyxl.load_workbook(arquivo)
            ws = wb.active

            # Encontra cabeçalho
            header_row = None
            for row in range(1, 20):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and str(cell_value).lower() in ['cpf', 'cnpj', 'cpf_cnpj', 'documento']:
                    header_row = row
                    break

            if not header_row:
                return {
                    'sucesso': False,
                    'erro': 'Cabeçalho não encontrado'
                }

            # Mapeia colunas
            headers = {}
            for col in range(1, ws.max_column + 1):
                header = str(ws.cell(row=header_row, column=col).value or '').lower().strip()
                headers[header] = col

            clientes = []
            erros = []

            for row in range(header_row + 1, ws.max_row + 1):
                cpf_cnpj = ws.cell(row=row, column=headers.get('cpf_cnpj', headers.get('cpf', headers.get('cnpj', 1)))).value
                if not cpf_cnpj:
                    continue

                try:
                    cliente = {
                        'cpf_cnpj': str(cpf_cnpj).strip(),
                        'nome': str(ws.cell(row=row, column=headers.get('nome', 2)).value or '').strip(),
                        'email': str(ws.cell(row=row, column=headers.get('email', 3)).value or '').strip(),
                        'telefone': str(ws.cell(row=row, column=headers.get('telefone', 4)).value or '').strip(),
                        'cidade': str(ws.cell(row=row, column=headers.get('cidade', 5)).value or '').strip() if 'cidade' in headers else '',
                        'uf': str(ws.cell(row=row, column=headers.get('uf', headers.get('estado', 6))).value or '').strip().upper() if 'uf' in headers or 'estado' in headers else '',
                        'linha': row
                    }

                    if not cliente['nome']:
                        erros.append(f"Linha {row}: Nome é obrigatório")
                        continue

                    clientes.append(cliente)

                except Exception as e:
                    erros.append(f"Linha {row}: {str(e)}")

            return {
                'sucesso': True,
                'clientes': clientes,
                'total': len(clientes),
                'erros': erros
            }

        except Exception as e:
            logger.error(f"Erro ao importar clientes: {str(e)}")
            return {
                'sucesso': False,
                'erro': str(e)
            }

    def gerar_modelo_importacao(self, tipo):
        """
        Gera modelo Excel para importação

        Args:
            tipo: Tipo de modelo ('produtos', 'clientes', 'estoque')

        Returns:
            bytes: Arquivo Excel modelo
        """
        wb = openpyxl.Workbook()
        ws = wb.active

        if tipo == 'produtos':
            ws.title = 'Produtos'
            headers = ['Código*', 'Nome*', 'Preço*', 'Preço Custo', 'Preço Promo',
                       'Categoria', 'NCM', 'Unidade', 'Estoque', 'Descrição']
            exemplos = ['PROD001', 'Produto Exemplo', '99.90', '50.00', '',
                        'Categoria 1', '84819090', 'UN', '100', 'Descrição do produto']

        elif tipo == 'clientes':
            ws.title = 'Clientes'
            headers = ['CPF/CNPJ*', 'Nome*', 'Email', 'Telefone', 'Celular',
                       'Endereço', 'Cidade', 'UF', 'CEP', 'Tipo']
            exemplos = ['12345678901', 'Cliente Exemplo', 'email@exemplo.com', '(11) 1234-5678', '',
                        'Rua Exemplo, 123', 'São Paulo', 'SP', '01234-567', 'varejo']

        elif tipo == 'estoque':
            ws.title = 'Estoque'
            headers = ['Código Produto*', 'Quantidade*', 'Localização', 'Lote',
                       'Validade', 'Mínimo', 'Máximo']
            exemplos = ['PROD001', '100', 'A1-01', 'LOTE001',
                        '31/12/2025', '10', '500']

        else:
            return None

        # Cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER_STYLE

        # Exemplo
        for col, valor in enumerate(exemplos, 1):
            cell = ws.cell(row=2, column=col, value=valor)
            cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

        # Instruções
        ws.cell(row=4, column=1, value='* Campos obrigatórios')
        ws.cell(row=5, column=1, value='A linha 2 contém um exemplo - substitua pelos seus dados')
        ws.cell(row=6, column=1, value='Não altere os cabeçalhos da linha 1')

        self._ajustar_larguras(ws, headers)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _adicionar_cabecalho_empresa(self, ws):
        """Adiciona cabeçalho com dados da empresa"""
        if self.empresa:
            ws.merge_cells('A1:J1')
            ws['A1'] = self.empresa.razao_social
            ws['A1'].font = Font(bold=True, size=16, color='0071E3')
            ws['A1'].alignment = Alignment(horizontal='center')

            ws.merge_cells('A2:J2')
            ws['A2'] = f'{self.empresa.logradouro}, {self.empresa.numero} - {self.empresa.cidade}/{self.empresa.uf}'
            ws['A2'].alignment = Alignment(horizontal='center')

            ws.merge_cells('A3:J3')
            ws['A3'] = f'CNPJ: {self.empresa.cnpj}'
            ws['A3'].alignment = Alignment(horizontal='center')

    def _ajustar_larguras(self, ws, headers):
        """Ajusta largura das colunas automaticamente"""
        for col, header in enumerate(headers, 1):
            col_letter = get_column_letter(col)
            max_length = len(str(header))

            for row in range(1, min(ws.max_row + 1, 100)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    def _parse_decimal(self, value):
        """Converte valor para decimal"""
        if value is None:
            return Decimal('0')

        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))

        value = str(value).strip()
        value = value.replace('R$', '').replace(' ', '').strip()

        # Formato brasileiro (1.234,56) ou internacional (1,234.56)
        if ',' in value and '.' in value:
            if value.rfind(',') > value.rfind('.'):
                value = value.replace('.', '').replace(',', '.')
            else:
                value = value.replace(',', '')
        elif ',' in value:
            value = value.replace(',', '.')

        try:
            return Decimal(value) if value else Decimal('0')
        except:
            return Decimal('0')
